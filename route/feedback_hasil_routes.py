from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, jsonify
import pandas as pd
from datetime import datetime
from function import feedbackHasils
from sklearn.cluster import DBSCAN
from sklearn.preprocessing import StandardScaler
from io import BytesIO
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from flask_mysqldb import MySQL

feedback_hasil_routes = Blueprint('feedback_hasil_routes', __name__)
mysql = MySQL()

@feedback_hasil_routes.route('/dashboard/feedback/hasil', methods = ['GET'])
def getFeedbackHasil():
    if 'username' in session:
        klustersPositif = feedbackHasils.fetch_klusters('Positif', session['user_id'])
        klustersNegatif = feedbackHasils.fetch_klusters('Negatif', session['user_id'])
        klustersNetral = feedbackHasils.fetch_klusters('Netral', session['user_id'])
        

        labels_doughnut_positif, values_doughnut_positif = feedbackHasils.get_top_programs(klustersPositif)
        labels_doughnut_negatif, values_doughnut_negatif = feedbackHasils.get_top_programs(klustersNegatif)
        labels_doughnut_netral, values_doughnut_netral = feedbackHasils.get_top_programs(klustersNetral)

        labels_bar_positif, values_bar_positif = feedbackHasils.get_top_batch(klustersPositif)
        labels_bar_negatif, values_bar_negatif = feedbackHasils.get_top_batch(klustersNegatif)
        labels_bar_netral, values_bar_netral = feedbackHasils.get_top_batch(klustersNetral)
        
        cur = mysql.connection.cursor()
        cur.execute(f"SELECT kluster, COUNT(*) AS jumlah_record FROM kluster WHERE kluster IN ('Positif', 'Negatif', 'Netral') AND user_id = { session['user_id']} GROUP BY kluster")
        grafis = cur.fetchall()
        mysql.connection.commit()
        cur.close()
        tabs = [
            {'id': 'positif', 'label': 'Positif', 'charts': ['myChart2', 'myChart3'], 'name': ['Program', 'Batch'], 'table_id': 'myTable1', 'klusters': klustersPositif},
            {'id': 'negatif', 'label': 'Negatif', 'charts': ['myChart4', 'myChart5'], 'name': ['Program', 'Batch'], 'table_id': 'myTable2', 'klusters': klustersNegatif},
            {'id': 'netral', 'label': 'Netral', 'charts': ['myChart6', 'myChart7'], 'name': ['Program', 'Batch'], 'table_id': 'myTable3', 'klusters': klustersNetral}
        ]

        labels = [item[0] for item in grafis]
        values = [item[1] for item in grafis]
        grafis = [
            labels,
            values
        ]

        silhouette_score = session.get('silhouette_score')
        category_choice = session.get('category_choice')
        
        return render_template('page/feedback/hasil.html', title='Hasil Klasterisasi Feedback', 
                               tabs=tabs, 
                               labels_doughnut_positif=labels_doughnut_positif, 
                               labels_doughnut_negatif=labels_doughnut_negatif, 
                               labels_doughnut_netral=labels_doughnut_netral, 
                               values_doughnut_positif=values_doughnut_positif, 
                               values_doughnut_negatif=values_doughnut_negatif, 
                               values_doughnut_netral=values_doughnut_netral, 
                               labels_bar_positif=labels_bar_positif, 
                               labels_bar_negatif=labels_bar_negatif, 
                               labels_bar_netral=labels_bar_netral, 
                               values_bar_positif=values_bar_positif,   
                               values_bar_negatif=values_bar_negatif, 
                               values_bar_netral=values_bar_netral, 
                               grafis=grafis,
                               silhouette_score=silhouette_score,
                               category_choice=category_choice)
    else:
        return redirect('/login')
    
@feedback_hasil_routes.route('/dashboard/feedback/hasil', methods=['POST'])
def postFeedbackHasil():
    # catch data feedback choice
    mappingChoice = {
        'Pembelajaran dan Pengajaran': 'pembelajaran_pengajaran',
        'Fasilitas dan Lingkungan': 'fasilitas_lingkungan',
        'Kepuasan terhadap Mentor': 'kepuasan_mentor'
    }

    choice = request.form['choice']
    for key, value in mappingChoice.items():
        choice = choice.replace(key, value)

    # hit database
    cur = mysql.connection.cursor()
    cur.execute(f"""
        SELECT tf.feedback_id, tf.translate_{choice}, fb.{choice} 
        FROM translate_feedback tf JOIN feedback fb ON tf.feedback_id = fb.id""")
    data = cur.fetchall()
    mysql.connection.commit()
    cur.close()

    # check data empty
    if len(data) == 0:
        flash('Data Feedback Kosong, Gagal Di Kluster', 'danger')
        return redirect(url_for('feedback_hasil_routes.postFeedbackHasil'))
    else:
        # convert tuple to dictionary form access by row now access by coloumn
        dictionary = {
                'id': [subtuple[0] for subtuple in data],
                'translated': [subtuple[1] for subtuple in data],
                'text': [subtuple[2] for subtuple in data],
        }

        # convert dataframe dan delete row when nan or empty string
        data = pd.DataFrame(dictionary)
        data.replace('', pd.NA, inplace=True)
        data = data.dropna(how='any')

        # prepossesing
        data['case folding'] = data['translated'].apply(feedbackHasils.case_folding)
        data['cleaning data'] = data['case folding'].apply(feedbackHasils.clean_text)
        data['tokenize'] = data['cleaning data'].apply(feedbackHasils.tokenize_text)
        data['stopwords'] = data['tokenize'].apply(feedbackHasils.remove_stopwords)
        data['stemming'] = data['stopwords'].apply(feedbackHasils.stem_tokens)

        # scoring sentiment
        data['polarity'] = data['stemming'].apply(feedbackHasils.getPolarity)
        data['subjectivity'] = data['stemming'].apply(feedbackHasils.getSubjectivity)
        data['afinn'] = data['stemming'].apply(feedbackHasils.getAfinnScore)

        # Standarisasi data
        scaler = StandardScaler()
        data_selected = data[['polarity', 'subjectivity', 'afinn']]
        data_scaled = scaler.fit_transform(data_selected)

        best_eps, best_min_samples, best_score = feedbackHasils.chekEpsMinRadius(data_scaled)

        session['category_choice'] = request.form['choice']
        print(best_score)
        
        # Menentukan parameter DBSCAN
        eps = best_eps
        min_samples = best_min_samples

        # Membuat model DBSCAN
        dbscan = DBSCAN(eps=eps, min_samples=min_samples, metric='euclidean')

        # Melakukan clusteringc
        clusters = dbscan.fit_predict(data_scaled)

        # Menambahkan kolom cluster ke DataFrame
        data['cluster'] = clusters
        
        data['sentiment'] = data['cluster'].apply(feedbackHasils.get_sentiment_label)
        grafis = pd.DataFrame()   
        grafis = data[['polarity','subjectivity','afinn','sentiment']]
        session['plot'] = grafis.to_dict(orient='records')

        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        updated_at = created_at
        cur = mysql.connection.cursor()
        cur.execute("""DELETE FROM kluster WHERE user_id=%s""", (session['user_id'],))
        for index, row in data.iterrows():
            cur.execute("INSERT INTO kluster (kluster,feedback, user_id, feedback_id, createdAt, updatedAt) VALUES (%s, %s, %s, %s, %s, %s)", 
                        (
                            row['sentiment'], 
                            row['text'], 
                            session['user_id'], 
                            row['id'], 
                            created_at,
                            updated_at
                            ))
        mysql.connection.commit()
        cur.close()

        flash('Data Berhasil Di Kluster', 'success')
        return redirect(url_for('feedback_hasil_routes.postFeedbackHasil'))

@feedback_hasil_routes.route('/dashboard/feedback/hasil/plot1_png')
def plot1_png():
    plot = session.get('plot')
    data = pd.DataFrame(plot)

    fig, ax = plt.subplots()

    for index, row in data.iterrows():
        if row['sentiment'] == 'Positif':
            ax.scatter(row['polarity'], row['subjectivity'], color="black", label='Positif')
        elif row['sentiment'] == 'Netral':
            ax.scatter(row['polarity'], row['subjectivity'], color="yellow", label='Netral')
        elif row['sentiment'] == 'Negatif':
            ax.scatter(row['polarity'], row['subjectivity'], color="red", label='Negatif')

    ax.set_title('DBSCAN Polarity and Subjectivity')
    ax.set_xlabel('polarity')
    ax.set_ylabel('subjectivity')

    # Tambahkan legenda
    handles, labels = ax.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax.legend(by_label.values(), by_label.keys())

    output = BytesIO()
    fig.savefig(output, format='png')
    output.seek(0)

    return send_file(output, mimetype='image/png')
@feedback_hasil_routes.route('/dashboard/feedback/hasil/plot2_png')
def plot2_png():
    plot = session.get('plot')
    data = pd.DataFrame(plot)

    fig, ax = plt.subplots()

    for index, row in data.iterrows():
        if row['sentiment'] == 'Positif':
            ax.scatter(row['polarity'], row['afinn'], color="black", label='Positif')
        elif row['sentiment'] == 'Netral':
            ax.scatter(row['polarity'], row['afinn'], color="yellow", label='Netral')
        elif row['sentiment'] == 'Negatif':
            ax.scatter(row['polarity'], row['afinn'], color="red", label='Negatif')

    ax.set_title('DBSCAN Polarity and Afinn Score')
    ax.set_xlabel('polarity')
    ax.set_ylabel('afinn')

    # Tambahkan legenda
    handles, labels = ax.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax.legend(by_label.values(), by_label.keys())

    output = BytesIO()
    fig.savefig(output, format='png')
    output.seek(0)

    return send_file(output, mimetype='image/png')

@feedback_hasil_routes.route('/dashboard/feedback/hasil/plot3_png')
def plot3_png():
    plot = session.get('plot')
    data = pd.DataFrame(plot)

    fig, ax = plt.subplots()

    for index, row in data.iterrows():
        if row['sentiment'] == 'Positif':
            ax.scatter(row['subjectivity'], row['afinn'], color="black", label='Positif')
        elif row['sentiment'] == 'Netral':
            ax.scatter(row['subjectivity'], row['afinn'], color="yellow", label='Netral')
        elif row['sentiment'] == 'Negatif':
            ax.scatter(row['subjectivity'], row['afinn'], color="red", label='Negatif')

    ax.set_title('DBSCAN Subjectivity and Afinn Score')
    ax.set_xlabel('subjectivity')
    ax.set_ylabel('afinn')

    # Tambahkan legenda
    handles, labels = ax.get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    ax.legend(by_label.values(), by_label.keys())

    output = BytesIO()
    fig.savefig(output, format='png')
    output.seek(0)

    return send_file(output, mimetype='image/png')

@feedback_hasil_routes.route('/dashboard/feedback/hasil/export_excel')
def export_excel():
    try:
        # Ambil data feedback
        feedback_data = feedbackHasils.get_feedback_kluster_from_database(session['user_id'])
        df = pd.DataFrame(feedback_data, columns=['kluster', 'feedback', 'feedback_name', 'jalur_pembelajaran', 'sesi', 'program', 'batch_name'])
        df.insert(0, 'no_urut', range(1, len(df) + 1))
        ordered_columns = ['no_urut', 'feedback_name', 'jalur_pembelajaran', 'program', 'batch_name', 'sesi', 'feedback', 'kluster']
        df = df[ordered_columns]
        df.columns = ['No Urut', 'Name', 'Jalur Pembelajaran', 'Program Name', 'Batch', 'Sesi', 'Feedback', 'Sentiment']

        # Buat workbook dan worksheet
        wb = Workbook()
        ws = wb.active

        # Set header style
        header_font = Font(name='Times New Roman', size=12, bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        # Isi data
        data_font = Font(name='Times New Roman', size=12)
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column width
        column_widths = {'A': 10, 'B': 50, 'C': 25, 'D': 50, 'E': 15, 'F': 15, 'G': 50, 'H': 50}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Set row height
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 70

        # Wrap text untuk kolom tertentu
        wrap_columns = ['G']
        for col in wrap_columns:
            for cell in ws[col]:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Fungsi untuk membuat grafik
        def create_graph(labels, values, title, graph_type='pie'):
            fig, ax = plt.subplots()
            if graph_type == 'pie' or graph_type == 'donut':
                wedges, texts, autotexts = ax.pie(
                    values, 
                    labels=labels, 
                    autopct='%1.1f%%', 
                    startangle=140,
                    wedgeprops=dict(width=0.55) if graph_type == 'donut' else None
                )
                # Mengatur warna teks
                for autotext in autotexts:
                    autotext.set_color('white')
                ax.set_title(title)
            elif graph_type == 'bar':
                ax.bar(labels, values)
                ax.set_title(title)
            plt.tight_layout()
            image = BytesIO()
            plt.savefig(image, format='png', bbox_inches='tight')
            image.seek(0)
            return ExcelImage(image)

        # Dapatkan data untuk grafik
        klustersPositif = feedbackHasils.fetch_klusters('Positif', session['user_id'])
        klustersNegatif = feedbackHasils.fetch_klusters('Negatif', session['user_id'])
        klustersNetral = feedbackHasils.fetch_klusters('Netral', session['user_id'])

        labels_doughnut_positif, values_doughnut_positif = feedbackHasils.get_top_programs(klustersPositif)
        labels_doughnut_negatif, values_doughnut_negatif = feedbackHasils.get_top_programs(klustersNegatif)
        labels_doughnut_netral, values_doughnut_netral = feedbackHasils.get_top_programs(klustersNetral)

        labels_bar_positif, values_bar_positif = feedbackHasils.get_top_batch(klustersPositif)
        labels_bar_negatif, values_bar_negatif = feedbackHasils.get_top_batch(klustersNegatif)
        labels_bar_netral, values_bar_netral = feedbackHasils.get_top_batch(klustersNetral)

        # Grafik pie untuk seluruh data
        cur = mysql.connection.cursor()
        cur.execute(f"SELECT kluster, COUNT(*) AS jumlah_record FROM kluster WHERE kluster IN ('Positif', 'Negatif', 'Netral') AND user_id = { session['user_id']} GROUP BY kluster")
        grafis = cur.fetchall()
        labels = [item[0] for item in grafis]
        values = [item[1] for item in grafis]
        grafis_pie_all = [labels, values]

        # Buat grafik dan tambahkan ke worksheet dengan jarak yang diatur
        def add_graph(ws, img, start_cell):
            ws.add_image(img, start_cell)

        # Tambahkan grafik ke worksheet
        add_graph(ws, create_graph(labels_doughnut_positif, values_doughnut_positif, 'Top Programs (Positif)', 'donut'), 'J2')
        add_graph(ws, create_graph(labels_doughnut_negatif, values_doughnut_negatif, 'Top Programs (Negatif)', 'donut'), 'J8')
        add_graph(ws, create_graph(labels_doughnut_netral, values_doughnut_netral, 'Top Programs (Netral)', 'donut'), 'J14')

        add_graph(ws, create_graph(labels_bar_positif, values_bar_positif, 'Top Batch (Positif)', 'bar'), 'X2')
        add_graph(ws, create_graph(labels_bar_negatif, values_bar_negatif, 'Top Batch (Negatif)', 'bar'), 'X8')
        add_graph(ws, create_graph(labels_bar_netral, values_bar_netral, 'Top Batch (Netral)', 'bar'), 'X14')

        add_graph(ws, create_graph(grafis_pie_all[0], grafis_pie_all[1], 'Overall Sentiment Distribution', 'pie'), 'J20')

        # Proses data dan buat grafik tambahan
        plot = session.get('plot')
        data = pd.DataFrame(plot)

        def create_scatter_plot(x_col, y_col, x_label, y_label, title):
            fig, ax = plt.subplots()
            for index, row in data.iterrows():
                if row['sentiment'] == 'Positif':
                    ax.scatter(row[x_col], row[y_col], color="black", label='Positif')
                elif row['sentiment'] == 'Netral':
                    ax.scatter(row[x_col], row[y_col], color="yellow", label='Netral')
                elif row['sentiment'] == 'Negatif':
                    ax.scatter(row[x_col], row[y_col], color="red", label='Negatif')
            ax.set_title(title)
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            handles, labels = ax.get_legend_handles_labels()
            by_label = dict(zip(labels, handles))
            ax.legend(by_label.values(), by_label.keys())
            image = BytesIO()
            fig.savefig(image, format='png')
            image.seek(0)
            return ExcelImage(image)

        add_graph(ws, create_scatter_plot('polarity', 'subjectivity', 'Polarity', 'Subjectivity', 'DBSCAN Polarity and Subjectivity'), 'X20')
        add_graph(ws, create_scatter_plot('polarity', 'afinn', 'Polarity', 'Afinn', 'DBSCAN Polarity and Afinn Score'), 'J26')
        add_graph(ws, create_scatter_plot('subjectivity', 'afinn', 'Subjectivity', 'Afinn', 'DBSCAN Subjectivity and Afinn Score'), 'X26')

        # Simpan file Excel
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(excel_file, download_name='feedback_kluster.xlsx', as_attachment=True)
    except Exception as e:
        flash(f'Excel Gagal Dicetak: {e}', 'danger')
        return redirect(url_for('feedback_hasil_routes.postFeedbackHasil'))