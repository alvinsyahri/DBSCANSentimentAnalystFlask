from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from function import feedbackDatas
from datetime import datetime
from flask_mysqldb import MySQL
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


feedback_data_routes = Blueprint('feedback_data_routes', __name__)
mysql = MySQL()

@feedback_data_routes.route('/dashboard/feedback/data', methods = ['GET'])
def getFeedbackData():
    if 'username' in session:
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT 
            f.id, 
            f.name, 
            f.jalur_pembelajaran, 
            (SELECT p.name FROM program p WHERE p.id = f.program_id) AS program_name,
            (SELECT b.name FROM batch b WHERE b.id = f.batch_id) AS batch_name,
            f.sesi, 
            f.pembelajaran_pengajaran, 
            f.fasilitas_lingkungan, 
            f.kepuasan_mentor
        FROM 
            feedback f
        ORDER BY 
            f.createdAt DESC;
        """)
        feedbacks = cur.fetchall()
        mysql.connection.commit()
        cur.close()
        return render_template('page/feedback/data.html', title='Data Feedback', feedbacks=feedbacks)
    else:
        return redirect('/login')  

@feedback_data_routes.route('/dashboard/feedback/data/export_excel')
def export_excel():
    try:
        feedback_data = feedbackDatas.get_feedback_data_from_database()

        df = pd.DataFrame(feedback_data, columns=['ID', 'Name', 'Jalur Pembelajaran', 'Program', 'Batch', 'Sesi', 'Pembelajaran dan Pembahasan', 'Fasilitas dan Lingkungan', 'Kepuasan terhadap Mentor'])
        df.rename(columns={'ID': 'No Urut'}, inplace=True)
        df['No Urut'] = range(1, len(df) + 1)

        wb = Workbook()
        ws = wb.active

        header_font = Font(name='Times New Roman', size=12, bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        data_font = Font(name='Times New Roman', size=12)
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 10  
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 50

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 70

        wrap_columns = ['G', 'H', 'I']
        for col in wrap_columns:
            for cell in ws[col]:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(excel_file, download_name='feedback_data.xlsx', as_attachment=True)
    except:
        flash('Excel Gagal Di Download', 'danger')
        return redirect(url_for('feedback_data_routes.getFeedbackData'))

@feedback_data_routes.route('/dashboard/feedback/data/<feedback_id>', methods=['POST', 'DELETE'])
def putAndDeleteBatch(feedback_id):
    if 'username' in session:
        if(request.form['_method'] == 'DELETE'):
            try:
                cur = mysql.connection.cursor()
                cur.execute("""DELETE FROM feedback WHERE id=%s""", (feedback_id))
                
                mysql.connection.commit()
                cur.close()

                flash('Feedback Berhasil Dihapus', 'success')
                return redirect(url_for('feedback_data_routes.getFeedbackData'))
            except:
                flash('Feedback Gagal Dihapus', 'danger')
                return redirect(url_for('feedback_data_routes.getFeedbackData'))
    else:
        return redirect('/login')
    
@feedback_data_routes.route('/dashboard/feedback/data', methods=['POST'])
def postDataFeedbackData():
    try:
    
        data = request.files['file']
        data = pd.read_excel(data)
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        updated_at = created_at
        convertProgram = feedbackDatas.convertProgram()
        convertBatch = feedbackDatas.convertBatch()
        cur = mysql.connection.cursor()
        # Iterasi melalui setiap baris DataFrame dan menyimpan data ke database
        for index, row in data.iterrows():
            # Menjalankan kueri untuk menyimpan data
            cur.execute("INSERT INTO feedback (name, jalur_pembelajaran, program_id, batch_id, sesi, pembelajaran_pengajaran, fasilitas_lingkungan, kepuasan_mentor, createdAt, updatedAt) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", 
                        (
                            row['Nama'], 
                            row['Jalur Pembelajaran'], 
                            convertProgram[row['Program']], 
                            convertBatch[row['Batch']], 
                            row['Sesi'],
                            row['Pembelajaran dan Pengajaran'],
                            row['Fasilitas dan Lingkungan'],
                            row['Kepuasan terhadap Mentor'],
                            created_at,
                            updated_at
                            ))
            # Ambil ID dari feedback yang baru saja dimasukkan
            feedback_id = cur.lastrowid

            # Terjemahkan teks untuk setiap kolom terkait
            translated_columns = {
                'translate_pembelajaran_pengajaran': row['Pembelajaran dan Pengajaran'],
                'translate_fasilitas_lingkungan': row['Fasilitas dan Lingkungan'],
                'translate_kepuasan_mentor': row['Kepuasan terhadap Mentor']
            }

            # Mengumpulkan hasil terjemahan
            translated_texts = {}
            for column, text in translated_columns.items():
                translated_texts[column] = feedbackDatas.translate_text(text)

            # Menjalankan kueri untuk menyimpan data ke tabel translate_feedback
            cur.execute(
                """
                INSERT INTO translate_feedback (feedback_id, translate_pembelajaran_pengajaran, translate_fasilitas_lingkungan, translate_kepuasan_mentor, createdAt, updatedAt) 
                VALUES (%s, %s, %s, %s, %s, %s)
                """, 
                (
                    feedback_id,
                    translated_texts['translate_pembelajaran_pengajaran'],
                    translated_texts['translate_fasilitas_lingkungan'],
                    translated_texts['translate_kepuasan_mentor'],
                    created_at,
                    updated_at
                )
            )

        mysql.connection.commit()
        cur.close()

        flash('Feedback Berhasil Di Import', 'success')
        return redirect(url_for('feedback_data_routes.getFeedbackData'))
    except Exception as e:
        flash(f'Feedback Gagal Di Import {e}', 'danger')
        return redirect(url_for('feedback_data_routes.getFeedbackData'))